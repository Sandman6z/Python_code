# -*- coding:utf-8

import openpyxl
import re

# 1. 数据指向excel中的第一个sheet
file_path = R"‪C:\Users\zbn02\Desktop\*.xls"
workbook = openpyxl.load_workbook(file_path)
sheet_names = workbook.sheetnames
sheet1 = workbook[sheet_names[0]]

# 2. 读取Excel sheet1中的所有数据
sheet1_allDatas = []
for row in sheet1.rows:
    line = [cell.value for cell in row]
    sheet1_allDatas.append(line)
#print(sheet1_allDatas)


# 3. 处理数据中姓名这个元素把字符串跟数字区分出来
# [1] 第一正则表达式
parrten1 = re.compile("[\d]+") #获取第一个元素的所有数字
parrten2 = re.compile("[^\d]+")#获取第一个元素的所有除数字以外的字符

# [2] 通过正则把元素的数字跟字符分离出来,并把处理好的数据添加到新的集合中
new_data = []
flg = 0
for strs in sheet1_allDatas:
    flg += 1
    for i in range(len(strs)):
        file_data1 = parrten1.findall(strs[0])
        file_data2 = parrten2.findall(strs[0])
        strs[0] = "".join(file_data2)
        strs.insert(1, "".join(file_data1))
    new_data.append(strs)
print(new_data)
print("====================================================")

# [3] 给title加电话标签
new_data[0].insert(3,"电话")
print(new_data)

# [4] 除去空格
last_Data = []
for lines in new_data:
    flg +=1
    for i in range(len(lines)):
        new_list =[x.strip() for x in lines if x.strip() != '']
    last_Data.append(new_list)
print(last_Data)

# [5] 创建新的Excel把结果写入到Excel中
wb = openpyxl.Workbook()
ws = workbook.active
ws.title = "sheet1"
flg = 0
for lines in last_Data:
    flg +=1
    for i in range(len(lines)):
        ws.cell(flg,i+1,lines[i])
    workbook.save("‪C:\Users\zbn02\Desktop\1.xlsx")
print("文件保存成功！")

